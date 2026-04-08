import asyncio
import io
import os
from datetime import datetime, timedelta

from fastapi import FastAPI, Query, HTTPException
from fastapi.responses import HTMLResponse, RedirectResponse, StreamingResponse
from fastapi.staticfiles import StaticFiles
from fastapi.templating import Jinja2Templates
from starlette.requests import Request

from app.clients.bling import BlingClient
from app.clients.mercadolivre import MercadoLivreClient
from app.services.sync_products import compare_products
from app.services.sync_orders import compare_orders
from app.services.sync_stock import compare_stock

BASE_PATH = os.environ.get("BASE_PATH", "")

app = FastAPI(title="Bling x ML Sync Dashboard")
app.mount("/static", StaticFiles(directory="app/static"), name="static")
templates = Jinja2Templates(directory="app/templates")
templates.env.globals["BASE"] = BASE_PATH


def default_dates():
    today = datetime.now().strftime("%Y-%m-%d")
    week_ago = (datetime.now() - timedelta(days=7)).strftime("%Y-%m-%d")
    return week_ago, today


# --- Pages ---

@app.get("/", response_class=HTMLResponse)
async def dashboard(request: Request):
    return templates.TemplateResponse(request, "dashboard.html")


@app.get("/bling", response_class=HTMLResponse)
async def bling_page(request: Request):
    return templates.TemplateResponse(request, "bling.html")


@app.get("/ml", response_class=HTMLResponse)
async def ml_page(request: Request):
    return templates.TemplateResponse(request, "ml.html")


# --- Bling OAuth2 ---

@app.get("/bling/auth")
async def bling_auth():
    bling = BlingClient()
    return RedirectResponse(bling.get_authorize_url())


@app.get("/bling/callback")
async def bling_callback(
    code: str = Query(None),
    error: str = Query(None),
    error_description: str = Query(None),
):
    if error:
        return {
            "status": "error",
            "error": error,
            "detail": error_description or error,
        }
    if not code:
        return {"status": "error", "detail": "Nenhum code recebido do Bling."}
    bling = BlingClient()
    try:
        await bling.exchange_code(code)
        return RedirectResponse(f"{BASE_PATH}/bling")
    except Exception as e:
        return {"status": "error", "detail": f"Erro ao trocar code por token: {e}"}


# --- Bling Independent API ---

@app.get("/api/bling/summary")
async def bling_summary(
    date_from: str = Query(None),
    date_to: str = Query(None),
):
    d_from, d_to = default_dates()
    date_from = date_from or d_from
    date_to = date_to or d_to

    bling = BlingClient()
    if not bling.access_token:
        raise HTTPException(
            status_code=401,
            detail="Bling nao autorizado. Acesse /bling/auth para conectar sua conta.",
        )
    try:
        products, orders = await asyncio.gather(
            bling.get_products(),
            bling.get_orders(date_from, date_to),
        )
        return {
            "products": [p.model_dump() for p in products],
            "orders": [o.model_dump() for o in orders],
        }
    finally:
        await bling.close()


@app.get("/bling/pedidos", response_class=HTMLResponse)
async def bling_orders_page(request: Request):
    return templates.TemplateResponse(request, "bling_orders.html")


@app.get("/bling/pedidos/{order_id}", response_class=HTMLResponse)
async def bling_order_detail_page(request: Request, order_id: int):
    return templates.TemplateResponse(request, "bling_order_detail.html", {"order_id": order_id})


@app.get("/api/bling/orders")
async def bling_orders_api(
    date_from: str = Query(None),
    date_to: str = Query(None),
):
    d_from, d_to = default_dates()
    date_from = date_from or d_from
    date_to = date_to or d_to

    bling = BlingClient()
    if not bling.access_token:
        raise HTTPException(status_code=401, detail="Bling nao autorizado.")
    try:
        orders, nfes = await asyncio.gather(
            bling.get_orders_raw(date_from, date_to),
            bling.get_nfe_list(),
        )

        # Build lookup: contato.id -> best NFe (prefer authorized)
        # NFe situacao: 1=Pendente, 2=Rejeitada, 3=Denegada, 4=Cancelada, 5=Autorizada, 6=Inutilizada
        nfe_by_contato = {}
        for nfe in nfes:
            cid = nfe.get("contato", {}).get("id", 0)
            if not cid:
                continue
            existing = nfe_by_contato.get(cid)
            if not existing or nfe.get("situacao", 0) > existing.get("situacao", 0):
                nfe_by_contato[cid] = nfe

        for order in orders:
            cid = order.get("contato", {}).get("id", 0)
            nfe = nfe_by_contato.get(cid)
            if nfe:
                order["nfe"] = {
                    "id": nfe["id"],
                    "numero": nfe.get("numero", ""),
                    "situacao": nfe.get("situacao", 0),
                    "chaveAcesso": nfe.get("chaveAcesso", ""),
                }
            else:
                order["nfe"] = None

        return orders
    finally:
        await bling.close()


@app.get("/api/bling/canais-venda")
async def bling_canais_venda():
    bling = BlingClient()
    if not bling.access_token:
        raise HTTPException(status_code=401, detail="Bling nao autorizado.")
    try:
        data = await bling._request("GET", "/canais-venda")
        return data.get("data", [])
    finally:
        await bling.close()


@app.get("/api/bling/produtos/export")
async def bling_export_products(canal: int = Query(None)):
    """Export Bling products to Excel. Optionally filter by canal de venda."""
    import httpx
    from openpyxl import Workbook
    from openpyxl.drawing.image import Image as XlImage
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    bling = BlingClient()
    if not bling.access_token:
        raise HTTPException(status_code=401, detail="Bling nao autorizado.")
    try:
        products = await bling.get_products_raw()
        canal_name = "todos"

        if canal:
            # Get canal info for the filename
            canais = (await bling._request("GET", "/canais-venda")).get("data", [])
            canal_info = next((c for c in canais if c["id"] == canal), None)
            if canal_info:
                canal_name = canal_info["descricao"].lower().replace(" ", "_")

            # Get all orders from this canal and extract product SKUs
            skus_no_canal = set()
            page = 1
            while True:
                data = await bling._request(
                    "GET", "/pedidos/vendas",
                    params={"pagina": page, "limite": 100},
                )
                orders = data.get("data", [])
                if not orders:
                    break
                for order in orders:
                    loja = order.get("loja", {})
                    if loja.get("id") == canal:
                        # Fetch order detail for items
                        detail = await bling._request(
                            "GET", f"/pedidos/vendas/{order['id']}"
                        )
                        order_data = detail.get("data", {})
                        for item in order_data.get("itens", []):
                            sku = item.get("codigo", "")
                            if sku:
                                skus_no_canal.add(sku)
                if len(orders) < 100:
                    break
                page += 1

            products = [p for p in products if p.get("codigo", "") in skus_no_canal]
    finally:
        await bling.close()

    wb = Workbook()
    ws = wb.active
    ws.title = "Produtos Bling"

    headers = [
        "Imagem", "Codigo (SKU)", "Nome", "Preco Venda", "Preco Custo",
        "Estoque", "Situacao", "Tipo", "Formato", "ID Bling",
    ]

    header_font = Font(bold=True, color="FFFFFF", size=10)
    header_fill = PatternFill(start_color="6B7A3D", end_color="6B7A3D", fill_type="solid")
    thin_border = Border(
        bottom=Side(style="thin", color="CCCCCC"),
    )

    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")

    sit_map = {"A": "Ativo", "I": "Inativo", "E": "Excluido"}
    tipo_map = {"P": "Produto", "S": "Servico"}
    fmt_map = {"S": "Simples", "V": "Com variacoes", "E": "Com composicao"}

    # Download images concurrently
    img_urls = [p.get("imagemURL", "") for p in products]
    img_data = {}
    async with httpx.AsyncClient(timeout=10.0) as http:
        sem = asyncio.Semaphore(20)

        async def fetch_img(url):
            if not url:
                return url, None
            async with sem:
                try:
                    r = await http.get(url)
                    if r.status_code == 200:
                        return url, r.content
                except Exception:
                    pass
                return url, None

        results = await asyncio.gather(*[fetch_img(u) for u in img_urls])
        img_data = {u: d for u, d in results}

    img_col_letter = "A"
    ws.column_dimensions[img_col_letter].width = 10

    for row_idx, p in enumerate(products, 2):
        est = p.get("estoque", {})
        img_url = p.get("imagemURL", "")

        values = [
            "",  # placeholder for image column
            p.get("codigo", ""),
            p.get("nome", ""),
            p.get("preco", 0),
            p.get("precoCusto", 0),
            est.get("saldoVirtualTotal", 0),
            sit_map.get(p.get("situacao", ""), p.get("situacao", "")),
            tipo_map.get(p.get("tipo", ""), p.get("tipo", "")),
            fmt_map.get(p.get("formato", ""), p.get("formato", "")),
            p.get("id", ""),
        ]

        for col, v in enumerate(values, 1):
            cell = ws.cell(row=row_idx, column=col, value=v)
            cell.border = thin_border

        ws.row_dimensions[row_idx].height = 50

        raw = img_data.get(img_url)
        if raw:
            try:
                img_buf = io.BytesIO(raw)
                img = XlImage(img_buf)
                img.width = 50
                img.height = 50
                ws.add_image(img, f"{img_col_letter}{row_idx}")
            except Exception:
                ws.cell(row=row_idx, column=1, value=img_url)

    col_widths = [10, 18, 40, 14, 14, 10, 10, 10, 16, 16]
    for col, w in enumerate(col_widths, 1):
        ws.column_dimensions[ws.cell(row=1, column=col).column_letter].width = w

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    filename = f"produtos_bling_{canal_name}_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"},
    )


@app.get("/api/bling/empresa")
async def bling_empresa():
    bling = BlingClient()
    if not bling.access_token:
        raise HTTPException(status_code=401, detail="Bling nao autorizado.")
    try:
        return await bling.get_empresa()
    except Exception as e:
        raise HTTPException(status_code=400, detail=str(e))
    finally:
        await bling.close()


@app.get("/api/bling/orders/{order_id}")
async def bling_order_detail_api(order_id: int):
    bling = BlingClient()
    if not bling.access_token:
        raise HTTPException(status_code=401, detail="Bling nao autorizado.")
    try:
        return await bling.get_order_detail(order_id)
    finally:
        await bling.close()


@app.get("/api/bling/orders/{order_id}/contas-receber")
async def bling_contas_receber(order_id: int):
    """Fetch accounts receivable linked to a sales order."""
    bling = BlingClient()
    if not bling.access_token:
        raise HTTPException(status_code=401, detail="Bling nao autorizado.")
    try:
        return await bling.get_contas_receber_by_origin(order_id)
    except Exception as e:
        raise HTTPException(status_code=400, detail=str(e))
    finally:
        await bling.close()


@app.get("/api/bling/orders/{order_id}/nfe")
async def bling_find_nfe(order_id: int):
    """Find existing NFe for an order (by notaFiscal.id or contato+valor match)."""
    bling = BlingClient()
    if not bling.access_token:
        raise HTTPException(status_code=401, detail="Bling nao autorizado.")
    try:
        order = await bling.get_order_detail(order_id)
        nfe_id = order.get("notaFiscal", {}).get("id", 0)
        if nfe_id and nfe_id != 0:
            return await bling.get_nfe(nfe_id)

        # Search by contato + value
        contato_id = order.get("contato", {}).get("id", 0)
        order_total = order.get("total", 0)
        nfes = await bling.get_nfe_list()
        for nfe in nfes:
            if (nfe.get("contato", {}).get("id", 0) == contato_id
                    and nfe.get("situacao", 0) in (1, 5)):
                nfe_detail = await bling.get_nfe(nfe["id"])
                if abs(nfe_detail.get("valorNota", 0) - order_total) < 0.01:
                    return nfe_detail
        return None
    finally:
        await bling.close()


@app.post("/api/bling/orders/{order_id}/nfe")
async def bling_generate_nfe(order_id: int):
    bling = BlingClient()
    if not bling.access_token:
        raise HTTPException(status_code=401, detail="Bling nao autorizado.")
    try:
        return await bling.generate_nfe(order_id)
    except Exception as e:
        raise HTTPException(status_code=400, detail=str(e))
    finally:
        await bling.close()


@app.post("/api/bling/nfe/{nfe_id}/retry")
async def bling_retry_nfe(nfe_id: int, order_id: int = Query(...)):
    """Fix a pending NFe with correct data and re-send to SEFAZ."""
    bling = BlingClient()
    if not bling.access_token:
        raise HTTPException(status_code=401, detail="Bling nao autorizado.")
    try:
        return await bling.retry_nfe(nfe_id, order_id)
    except Exception as e:
        raise HTTPException(status_code=400, detail=str(e))
    finally:
        await bling.close()


@app.post("/api/bling/nfe/{nfe_id}/cancel")
async def bling_cancel_nfe(nfe_id: int):
    bling = BlingClient()
    if not bling.access_token:
        raise HTTPException(status_code=401, detail="Bling nao autorizado.")
    try:
        return await bling.cancel_nfe(nfe_id)
    except Exception as e:
        raise HTTPException(status_code=400, detail=str(e))
    finally:
        await bling.close()


@app.get("/api/bling/nfe/{nfe_id}")
async def bling_get_nfe(nfe_id: int):
    bling = BlingClient()
    if not bling.access_token:
        raise HTTPException(status_code=401, detail="Bling nao autorizado.")
    try:
        return await bling.get_nfe(nfe_id)
    finally:
        await bling.close()


# --- Mercado Livre Independent API ---

@app.get("/api/ml/summary")
async def ml_summary(
    date_from: str = Query(None),
    date_to: str = Query(None),
):
    d_from, d_to = default_dates()
    date_from = date_from or d_from
    date_to = date_to or d_to

    ml = MercadoLivreClient()
    if not ml.access_token or not ml.seller_id:
        raise HTTPException(
            status_code=401,
            detail="Mercado Livre nao configurado. Preencha as credenciais no arquivo .env",
        )
    try:
        items, orders = await asyncio.gather(
            ml.get_items(),
            ml.get_orders(date_from, date_to),
        )
        return {
            "items": [i.model_dump() for i in items],
            "orders": [o.model_dump() for o in orders],
        }
    finally:
        await ml.close()


# --- Sync/Comparison API (Consolidated) ---

@app.get("/api/sync/products")
async def sync_products(
    date_from: str = Query(None),
    date_to: str = Query(None),
):
    bling = BlingClient()
    ml = MercadoLivreClient()

    try:
        bling_products, ml_products = await asyncio.gather(
            bling.get_products(),
            ml.get_items(),
        )
        return compare_products(bling_products, ml_products)
    finally:
        await asyncio.gather(bling.close(), ml.close())


@app.get("/api/sync/orders")
async def sync_orders(
    date_from: str = Query(None),
    date_to: str = Query(None),
):
    d_from, d_to = default_dates()
    date_from = date_from or d_from
    date_to = date_to or d_to

    bling = BlingClient()
    ml = MercadoLivreClient()

    try:
        bling_orders, ml_orders = await asyncio.gather(
            bling.get_orders(date_from, date_to),
            ml.get_orders(date_from, date_to),
        )
        return compare_orders(bling_orders, ml_orders)
    finally:
        await asyncio.gather(bling.close(), ml.close())


@app.get("/api/sync/stock")
async def sync_stock(
    date_from: str = Query(None),
    date_to: str = Query(None),
):
    bling = BlingClient()
    ml = MercadoLivreClient()

    try:
        bling_products, ml_products = await asyncio.gather(
            bling.get_products(),
            ml.get_items(),
        )

        bling_stock = {p.sku: p.stock for p in bling_products if p.sku}
        ml_stock = {p.sku: p.stock for p in ml_products if p.sku}
        bling_names = {p.sku: p.name for p in bling_products if p.sku}

        return compare_stock(bling_stock, ml_stock, bling_names)
    finally:
        await asyncio.gather(bling.close(), ml.close())


@app.get("/api/sync/summary")
async def sync_summary(
    date_from: str = Query(None),
    date_to: str = Query(None),
):
    d_from, d_to = default_dates()
    date_from = date_from or d_from
    date_to = date_to or d_to

    bling = BlingClient()
    ml = MercadoLivreClient()

    try:
        bling_products, ml_products, bling_orders, ml_orders = await asyncio.gather(
            bling.get_products(),
            ml.get_items(),
            bling.get_orders(date_from, date_to),
            ml.get_orders(date_from, date_to),
        )

        prod_result = compare_products(bling_products, ml_products)

        bling_stock = {p.sku: p.stock for p in bling_products if p.sku}
        ml_stock = {p.sku: p.stock for p in ml_products if p.sku}
        stock_result = compare_stock(bling_stock, ml_stock)

        order_result = compare_orders(bling_orders, ml_orders)

        total_items = (
            prod_result["synced"] + prod_result["divergent"]
            + order_result["synced"] + order_result["divergent"]
            + stock_result["synced"] + stock_result["divergent"]
        )
        total_synced = prod_result["synced"] + order_result["synced"] + stock_result["synced"]
        health = (total_synced / total_items * 100) if total_items > 0 else 100.0

        return {
            "total_products_bling": prod_result["total_bling"],
            "total_products_ml": prod_result["total_ml"],
            "products_synced": prod_result["synced"],
            "products_divergent": prod_result["divergent"],
            "products_missing_ml": prod_result["missing_ml"],
            "products_missing_bling": prod_result["missing_bling"],
            "total_orders_bling": order_result["total_bling"],
            "total_orders_ml": order_result["total_ml"],
            "orders_synced": order_result["synced"],
            "orders_divergent": order_result["divergent"],
            "stock_synced": stock_result["synced"],
            "stock_divergent": stock_result["divergent"],
            "health_percentage": round(health, 1),
            "last_check": datetime.now().isoformat(),
        }
    finally:
        await asyncio.gather(bling.close(), ml.close())
